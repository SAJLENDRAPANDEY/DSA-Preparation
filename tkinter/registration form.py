import tkinter
from tkinter import messagebox,ttk
# ttk theme toolkit is use for combo box means slidebar ke liye kai sare option select krne ke liye 
import openpyxl
from openpyxl import load_workbook
# meassage box is used for store data
root=tkinter.Tk()
root.geometry("300x400")
# configure is used for change the background color 
# root.configure(bg="#ADD8E6")
path=r"C:\yash_important data\tkinter_excel_practicce.xlsx"
A=openpyxl.load_workbook(path)
B=A["Registration"]
 
# add 4 row using code first we check and then appen
# if B.max_row==1 and B.cell(row=1,column=1).value is None:
#     B.append(['name','email','phone','dropdown'])
#     A.save(path)

# Specially 4 collumn add krna ho tb 

if B.cell(row=1,column=4).value is None:
    B.cell(row=1,column=4).value="Department"
    A.save(path)

if B.cell(row=1,column=5).value is None:
    B.cell(row=1,column=5).value="check_boxx"
    A.save(path)

def all_command():
    name=name_box.get()
    email=Email_box.get()
    phone=phone_box.get()
    dropdown=combo_box.get()
    # check_boxx=check_var.get()

    # message=f"Name:{name}\n Email:{email}\n Phone:{phone}"
    # messagebox.showinfo("Capture Data",message)

    # check all fields are filled
    if name and email and phone and dropdown and check_var:
        B.append([name,email,phone,dropdown,check_var])
        A.save(path)
        messagebox.showinfo("Data submitted","Data submitted")
    else:
        messagebox.showwarning("warning","enter all fields")


# enter name label
name_label=tkinter.Label(root,text="Enter name ")
name_label.pack(anchor='w',padx=25)
name_box=tkinter.Entry(root)
# anchor set the direction of box means east=e,west=w,south=s,north=n and pading means gives the space of box means left to right how much move
name_box.pack(anchor='w')

# Email 
name_Email=tkinter.Label(root,text="Enter Email ")
name_Email.pack(anchor='w',padx=25)
Email_box=tkinter.Entry(root)
Email_box.pack(anchor='w')

# phone stored 


# phone 
phone_label=tkinter.Label(root,text="Enter phone ")
phone_label.pack(anchor='w',padx=25)
phone_box=tkinter.Entry(root)
phone_box.pack(anchor='w')

# dropdown lits
choices=["cs","it","ai","ds"]
combo_box_label=tkinter.Label(root,text="Select choice")
combo_box_label.pack(anchor='w',padx=25)
combo_box=ttk.Combobox(root,values=choices)
combo_box.pack(anchor='w')

# create check noxcheck_var=tkinter.IntVar()
check_var=tkinter.IntVar() # 0 = not check , 1 =check
check_box=ttk.Checkbutton(text="Please check it",variable=check_var)
check_box.pack(anchor='w',padx=25,pady=10)


# Submit 
submit_button=tkinter.Button(root,text="Submit",command=all_command)
submit_button.pack(anchor='w',padx=30)



root.mainloop()
